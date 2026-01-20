"""Export functionality for business data and crawl results."""
from __future__ import annotations
import json
import csv
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path


RESULTS_DIR = "results"


def _ensure_results_directory() -> None:
    """Ensure the results directory exists."""
    if not os.path.exists(RESULTS_DIR):
        os.makedirs(RESULTS_DIR)


def _create_filename(base_name: str, extension: str, timestamp: bool = True) -> str:
    """Create a safe filename with optional timestamp."""
    if timestamp:
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{base_name}_{timestamp_str}.{extension}"
    else:
        filename = f"{base_name}.{extension}"
    
    # Remove invalid characters
    filename = "".join(c for c in filename if c.isalnum() or c in "._-")
    return os.path.join(RESULTS_DIR, filename)


def export_json(data: Dict[str, Any], filename: Optional[str] = None, 
                 timestamp: bool = True) -> str:
    """
    Export data to JSON file.
    
    Args:
        data: Dictionary to export
        filename: Optional custom filename (without extension)
        timestamp: Whether to add timestamp to filename
        
    Returns:
        Path to the created file
    """
    _ensure_results_directory()
    
    base_name = filename or "business_data"
    filepath = _create_filename(base_name, "json", timestamp)
    
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return filepath
    except Exception as e:
        raise IOError(f"Failed to export JSON file: {e}") from e


def export_csv(business_info: Dict[str, Any], filename: Optional[str] = None,
               timestamp: bool = True) -> str:
    """
    Export business info to CSV file.
    
    Args:
        business_info: Business information dictionary
        filename: Optional custom filename (without extension)
        timestamp: Whether to add timestamp to filename
        
    Returns:
        Path to the created file
    """
    _ensure_results_directory()
    
    base_name = filename or "business_data"
    filepath = _create_filename(base_name, "csv", timestamp)
    
    # Flatten business info for CSV
    row = {
        "Business Name": business_info.get("business_name", ""),
        "Rating": business_info.get("rating", ""),
        "Review Count": business_info.get("review_count", ""),
        "Address": "; ".join(business_info.get("addresses", [])) if business_info.get("addresses") else "",
        "Areas Served": "; ".join(business_info.get("areas_served", [])) if business_info.get("areas_served") else "",
        "Hours": business_info.get("hours", ""),
        "Phone": "; ".join(business_info.get("phones", [])) if business_info.get("phones") else "",
        "Quote URL": business_info.get("quote_url", ""),
        "Services": "; ".join(business_info.get("services", [])) if business_info.get("services") else "",
        "Has Coupon": "Yes" if business_info.get("has_coupon") else "No",
        "24/7 Emergency": "Yes" if business_info.get("has_emergency") else "No",
        "Licensed": "Yes" if business_info.get("is_licensed") else "No",
        "License Number": business_info.get("license_number", ""),
        "Founding Date": business_info.get("founding_date", ""),
        "Founders": "; ".join([
            f"{f.get('name', '')} ({f.get('job_title', '')})" if f.get('job_title') 
            else f.get('name', '') 
            for f in business_info.get("founders", [])
        ]) if business_info.get("founders") else "",
    }
    
    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=row.keys())
            writer.writeheader()
            writer.writerow(row)
        return filepath
    except Exception as e:
        raise IOError(f"Failed to export CSV file: {e}") from e


def export_excel(business_info: Dict[str, Any], filename: Optional[str] = None,
                 timestamp: bool = True) -> str:
    """
    Export business info to Excel file.
    
    Args:
        business_info: Business information dictionary
        filename: Optional custom filename (without extension)
        timestamp: Whether to add timestamp to filename
        
    Returns:
        Path to the created file
        
    Raises:
        ImportError: If openpyxl is not installed
        IOError: If file creation fails
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from None
    
    _ensure_results_directory()
    
    base_name = filename or "business_data"
    filepath = _create_filename(base_name, "xlsx", timestamp)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Business Information"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Prepare data
    data = [
        ["Field", "Value"],
        ["Business Name", business_info.get("business_name", "")],
        ["Rating", f"{business_info.get('rating', '')}/5" if business_info.get("rating") else ""],
        ["Review Count", business_info.get("review_count", "")],
        ["Address", "; ".join(business_info.get("addresses", [])) if business_info.get("addresses") else ""],
        ["Areas Served", "; ".join(business_info.get("areas_served", [])) if business_info.get("areas_served") else ""],
        ["Hours", business_info.get("hours", "")],
        ["Phone", "; ".join(business_info.get("phones", [])) if business_info.get("phones") else ""],
        ["Quote URL", business_info.get("quote_url", "")],
        ["Services", "; ".join(business_info.get("services", [])) if business_info.get("services") else ""],
        ["Has Coupon", "Yes" if business_info.get("has_coupon") else "No"],
        ["24/7 Emergency", "Yes" if business_info.get("has_emergency") else "No"],
        ["Licensed", "Yes" if business_info.get("is_licensed") else "No"],
        ["License Number", business_info.get("license_number", "")],
        ["Founding Date", business_info.get("founding_date", "")],
    ]
    
    # Add founders row if available
    if business_info.get("founders"):
        founders_str = "; ".join([
            f"{f.get('name', '')} ({f.get('job_title', '')})" if f.get('job_title')
            else f.get('name', '')
            for f in business_info.get("founders", [])
        ])
        data.append(["Founders", founders_str])
    
    # Write data to worksheet
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style header row
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    
    try:
        wb.save(filepath)
        return filepath
    except Exception as e:
        raise IOError(f"Failed to export Excel file: {e}") from e


def export_crawl_results(results: Dict[str, Any], filename: Optional[str] = None,
                         timestamp: bool = True) -> str:
    """
    Export crawl results to JSON file.
    
    Args:
        results: Crawl results dictionary from scrape_priority_paths
        filename: Optional custom filename (without extension)
        timestamp: Whether to add timestamp to filename
        
    Returns:
        Path to the created file
    """
    return export_json(results, filename or "crawl_results", timestamp)


def export_batch_results(results_list: List[Dict[str, Any]], filename: Optional[str] = None,
                         timestamp: bool = True) -> str:
    """
    Export multiple business results to JSON file.
    
    Args:
        results_list: List of business info dictionaries
        filename: Optional custom filename (without extension)
        timestamp: Whether to add timestamp to filename
        
    Returns:
        Path to the created file
    """
    data = {
        "total_businesses": len(results_list),
        "exported_at": datetime.now().isoformat(),
        "businesses": results_list
    }
    return export_json(data, filename or "batch_results", timestamp)
